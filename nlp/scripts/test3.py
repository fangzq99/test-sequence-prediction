import openpyxl
import nltk

nltk.download("stopwords")
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import stopwords


# Load the workbook
workbook = openpyxl.load_workbook("test_names_raw_database.xlsx")

# Select the desired worksheet
worksheet = workbook["Sheet1"]  # Replace 'Sheet1' with the actual sheet name

# Specify the column and row numbers
column_number = 3  # Replace with the desired column number
row_number = 2  # Replace with the desired row number

# Access the cell value
cell_value = worksheet.cell(row=row_number, column=column_number).value

# Iterate through each row in the column
for row in worksheet.iter_rows(min_row=1, min_col=column_number, max_col=column_number):
    stop_words = set(stopwords.words("english"))

    cell = row[0]
    cell_value = cell.value
    words_in_quote = word_tokenize(cell_value)
    filtered_list = []

    for word in words_in_quote:
        if word.casefold() not in stop_words:
            filtered_list.append(word)
    print(words_in_quote)
    print(filtered_list)
    print("")
